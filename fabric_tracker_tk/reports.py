import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fabric_tracker_tk import db
from datetime import datetime

COLS       = ("firm", "date", "batch", "supplier", "yarn", "kg", "rolls", "delivered")
HEADINGS   = ["Firm", "Date", "Batch", "Supplier", "Yarn", "Kg", "Rolls", "Delivered To"]
WIDTHS     = [140, 90, 90, 150, 150, 80, 80, 140]

class ReportsFrame(ttk.Frame):
    def __init__(self, parent, controller=None):
        super().__init__(parent)
        self.controller = controller
        self.build_ui()

    def build_ui(self):
        top = ttk.Frame(self)
        top.pack(fill="x", padx=6, pady=6)

        # --- Financial Year ---
        ttk.Label(top, text="Financial Year Start:").grid(row=0, column=0, sticky="w", padx=4)
        self.fy_start = ttk.Combobox(top, values=self._generate_fy_years(), width=8, state="readonly")
        self.fy_start.grid(row=0, column=1, padx=4)
        self.fy_start.set(self._default_fy())

        # --- Firm Filter ---
        ttk.Label(top, text="Firm:").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self._firm_var = tk.StringVar(value="Both (Combined)")
        firm_options = ["Both (Combined)"] + db.FIRMS
        self.firm_cb = ttk.Combobox(top, textvariable=self._firm_var,
                                    values=firm_options, state="readonly", width=22)
        self.firm_cb.grid(row=0, column=3, padx=4)

        # --- Buttons ---
        ttk.Button(top, text="Apply", command=self.load_report).grid(row=0, column=4, padx=6)
        ttk.Button(top, text="Export to Excel", command=self.export_report).grid(row=0, column=5, padx=4)

        # --- Summary bar ---
        self._summary_var = tk.StringVar(value="")
        ttk.Label(self, textvariable=self._summary_var, foreground="gray").pack(anchor="w", padx=8)

        # --- Table ---
        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, padx=6, pady=4)
        sy = ttk.Scrollbar(frame, orient="vertical")
        sy.pack(side="right", fill="y")
        sx = ttk.Scrollbar(frame, orient="horizontal")
        sx.pack(side="bottom", fill="x")
        self.tree = ttk.Treeview(frame, columns=COLS, show="headings",
                                  yscrollcommand=sy.set, xscrollcommand=sx.set)
        self.tree.pack(fill="both", expand=True)
        sy.config(command=self.tree.yview)
        sx.config(command=self.tree.xview)
        for c, h, w in zip(COLS, HEADINGS, WIDTHS):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w)

        # Tag rows by firm for colour coding
        self.tree.tag_configure(db.FIRMS[0], background="#e8f4fd")   # light blue — SP
        self.tree.tag_configure(db.FIRMS[1], background="#fdf6e8")   # light amber — RK

    # ------------------------------------------------------------------
    def _generate_fy_years(self):
        now = datetime.now().year
        return [str(y) for y in range(now - 5, now + 1)]

    def _default_fy(self):
        now = datetime.now()
        return str(now.year) if now.month >= 4 else str(now.year - 1)

    def _fy_dates(self):
        start_year = int(self.fy_start.get())
        return f"{start_year}-04-01", f"{start_year + 1}-03-31"

    def _fetch_rows(self, firm_name=None):
        """Return purchase rows for the selected FY, optionally filtered by firm."""
        fy_start, fy_end = self._fy_dates()
        with db.get_connection() as conn:
            cur = conn.cursor()
            if firm_name:
                cur.execute("""
                    SELECT firm_name, date, batch_id, supplier, yarn_type,
                           qty_kg, qty_rolls, delivered_to
                    FROM purchases
                    WHERE date >= ? AND date <= ? AND firm_name = ?
                    ORDER BY firm_name, date
                """, (fy_start, fy_end, firm_name))
            else:
                cur.execute("""
                    SELECT firm_name, date, batch_id, supplier, yarn_type,
                           qty_kg, qty_rolls, delivered_to
                    FROM purchases
                    WHERE date >= ? AND date <= ?
                    ORDER BY firm_name, date
                """, (fy_start, fy_end))
            return cur.fetchall()

    def load_report(self):
        for r in self.tree.get_children():
            self.tree.delete(r)

        selected = self._firm_var.get()
        firm_filter = None if selected == "Both (Combined)" else selected
        rows = self._fetch_rows(firm_filter)

        total_kg = 0
        total_rolls = 0
        for row in rows:
            display_date = db.db_to_ui_date(row["date"])
            firm = row["firm_name"] or ""
            kg = row["qty_kg"] or 0
            rolls = row["qty_rolls"] or 0
            total_kg += kg
            total_rolls += rolls
            self.tree.insert("", "end", values=(
                firm, display_date, row["batch_id"], row["supplier"],
                row["yarn_type"], kg, rolls, row["delivered_to"]
            ), tags=(firm,))

        self._summary_var.set(
            f"  {len(rows)} records  |  Total Kg: {total_kg:,.2f}  |  Total Rolls: {total_rolls:,}"
        )

    def export_report(self):
        selected = self._firm_var.get()
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"FabricReport_{self.fy_start.get()}.xlsx"
        )
        if not file_path:
            return

        wb = Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        header = ["Firm", "Date", "Batch", "Supplier", "Yarn", "Kg", "Rolls", "Delivered To"]
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="2E4057")

        def make_sheet(ws, rows, title_firm=""):
            ws.title = title_firm or "Combined"
            # Header
            ws.append(header)
            for cell in ws[1]:
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
            # Data
            firm_colours = {
                db.FIRMS[0]: "D6EAF8",   # blue tint
                db.FIRMS[1]: "FEF9E7",   # amber tint
            }
            for row in rows:
                display_date = db.db_to_ui_date(row["date"])
                firm = row["firm_name"] or ""
                ws.append([firm, display_date, row["batch_id"], row["supplier"],
                            row["yarn_type"], row["qty_kg"], row["qty_rolls"],
                            row["delivered_to"]])
                fill_color = firm_colours.get(firm, "FFFFFF")
                row_idx = ws.max_row
                for cell in ws[row_idx]:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
            # Totals row
            last = ws.max_row
            ws.append(["", "TOTAL", "", "", "",
                        f"=SUM(F2:F{last})", f"=SUM(G2:G{last})", ""])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            # Column widths
            for col, w in zip("ABCDEFGH", [24, 12, 14, 20, 20, 10, 10, 20]):
                ws.column_dimensions[col].width = w

        if selected == "Both (Combined)":
            # Three sheets: SP, RK, Combined
            for firm in db.FIRMS:
                rows = self._fetch_rows(firm)
                short = firm.split()[0]          # "S.P." or "R"
                ws = wb.create_sheet(title=firm[:28])
                make_sheet(ws, rows, firm)
            # Combined sheet
            all_rows = self._fetch_rows(None)
            ws_combined = wb.create_sheet(title="Combined")
            make_sheet(ws_combined, all_rows, "")
        else:
            rows = self._fetch_rows(selected)
            ws = wb.create_sheet(title=selected[:28])
            make_sheet(ws, rows, selected)

        wb.save(file_path)
        messagebox.showinfo("Export Complete", f"Report saved to:\n{file_path}")

    def reload_data(self):
        self.load_report()
